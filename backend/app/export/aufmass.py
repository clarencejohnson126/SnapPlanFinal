"""
Schicht 3 — turning reviewed positions into documents that leave the system.

THE EXPORT GATE

Everything here refuses to emit a position that no human has signed off on.
That is not a policy setting, it is the point: an unreviewed machine proposal
reaching an invoice is the one failure mode that would end the product.
`Position.is_exportable` is the single source of truth, and `_gate()` enforces
it before any document is built.

The Aufmaßprotokoll is the more important of the two outputs. Excel carries the
numbers; the protocol carries the *reasoning*, which is what a Prüfer asks for
when a quantity is disputed.
"""

from datetime import datetime, timezone
from io import BytesIO
from typing import Dict, List, Tuple

from app.domain.position import Position, PositionSet, format_de
from app.domain.trades import atv_for, label_for


class ExportBlockedError(RuntimeError):
    """
    Raised when an export is attempted with nothing signed off.

    Carries the counts so the caller can tell the user exactly what is missing
    rather than showing a generic failure.
    """

    def __init__(self, total: int, pending: int) -> None:
        self.total = total
        self.pending = pending
        super().__init__(
            f"Export nicht möglich: {pending} von {total} Positionen sind noch "
            f"nicht freigegeben. Bitte im Prüfmodus bestätigen."
        )


def _gate(position_set: PositionSet) -> List[Position]:
    """
    The exportable positions, or an error explaining why there are none.

    A partially reviewed set exports fine — the reviewed part. Only a set with
    nothing reviewed at all is refused, because that is the case where the user
    has misunderstood what they are about to send to a customer.
    """
    exportable = [p for p in position_set.positions if p.is_exportable]
    if not exportable:
        raise ExportBlockedError(
            total=position_set.total_count,
            pending=position_set.review_pending_count,
        )
    return exportable


def _header_lines(position_set: PositionSet, exportable: List[Position]) -> List[str]:
    """Shared document header for protocol and spreadsheet."""
    stamped = datetime.now(timezone.utc).astimezone().strftime("%d.%m.%Y %H:%M")
    lines = [
        f"Dokument:   {position_set.document_id}",
        f"Gewerk:     {label_for(position_set.trade)}",
        f"Grundlage:  VOB/C ATV {atv_for(position_set.trade)}",
        f"Regelwerk:  {position_set.ruleset_id} v{position_set.ruleset_version}",
        f"Erstellt:   {stamped}",
        f"Positionen: {len(exportable)} freigegeben von {position_set.total_count}",
    ]
    if position_set.review_pending_count:
        lines.append(
            f"HINWEIS:    {position_set.review_pending_count} Positionen sind "
            f"nicht freigegeben und in diesem Export NICHT enthalten."
        )
    return lines


def _totals(positions: List[Position]) -> List[Tuple[str, float]]:
    """Summed quantities per unit, in stable order."""
    totals: Dict[str, float] = {}
    for position in positions:
        totals[position.unit] = totals.get(position.unit, 0.0) + position.quantity
    return sorted(totals.items())


# ------------------------------------------------------------------- protocol

def build_protocol(position_set: PositionSet) -> str:
    """
    The Aufmaßprotokoll: every quantity with its full derivation.

    This is the document that answers "how did you get to 51,22 m²?". Each
    position shows its Rechenweg line by line, the norm citation that governed
    each deduction, who signed it off, and which page it came from.
    """
    exportable = _gate(position_set)

    out: List[str] = ["AUFMASSPROTOKOLL", "=" * 72]
    out.extend(_header_lines(position_set, exportable))
    out.append("=" * 72)
    out.append("")

    for index, position in enumerate(exportable, start=1):
        out.append(f"{index:>3}.  {position.designation}")

        for step in position.calculation:
            out.append(f"        {step.as_line()}")

        out.append(f"        ERGEBNIS: {format_de(position.quantity)} {position.unit}")

        if position.raw_quantity is not None and position.raw_quantity != position.quantity:
            out.append(
                f"        (geometrisch {format_de(position.raw_quantity)} "
                f"{position.unit} vor Anwendung der Abrechnungsregeln)"
            )

        pages = sorted({e.page_number for e in position.evidence})
        page_label = "Seite" if len(pages) == 1 else "Seiten"
        out.append(f"        Beleg: {page_label} {', '.join(str(p) for p in pages)}")

        if position.reviewed_by:
            out.append(
                f"        Freigabe: {position.reviewed_by} ({position.status.value})"
            )

        for warning in position.warnings:
            out.append(f"        ! {warning}")

        out.append("")

    out.append("-" * 72)
    out.append("SUMMEN (nur freigegebene Positionen)")
    for unit, total in _totals(exportable):
        out.append(f"    {format_de(total):>14} {unit}")

    return "\n".join(out)


# ---------------------------------------------------------------------- excel

#: Column layout. Kept here so protocol and spreadsheet stay recognisably the
#: same document in two forms.
_COLUMNS: List[Tuple[str, int]] = [
    ("Pos.", 6),
    ("Raum", 22),
    ("Bezeichnung", 38),
    ("Menge", 12),
    ("Einheit", 9),
    ("Roh (geom.)", 13),
    ("Rechenweg", 70),
    ("ATV", 12),
    ("Seite", 7),
    ("Status", 12),
    ("Freigabe", 16),
    ("Hinweise", 46),
]


def _position_row(index: int, position: Position, formatted: bool) -> List:
    """One export row. `formatted` picks German strings over raw floats."""
    pages = sorted({e.page_number for e in position.evidence})
    quantity = format_de(position.quantity) if formatted else round(position.quantity, 3)

    if position.raw_quantity is None:
        raw = ""
    elif formatted:
        raw = format_de(position.raw_quantity)
    else:
        raw = round(position.raw_quantity, 3)

    return [
        index,
        position.room_label or "",
        position.designation,
        quantity,
        position.unit,
        raw,
        " | ".join(position.protocol_lines()),
        position.atv,
        ", ".join(str(p) for p in pages),
        position.status.value,
        position.reviewed_by or "",
        " | ".join(position.warnings),
    ]


def build_excel(position_set: PositionSet) -> bytes:
    """
    Spreadsheet export, Aufmaß-ready.

    Carries the Rechenweg in its own column rather than only the result — an
    Excel row saying "51,22" and nothing else is not checkable, and a quantity
    nobody can check is a quantity nobody will pay for.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    exportable = _gate(position_set)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = label_for(position_set.trade)[:31]  # Excel caps sheet names

    bold = Font(bold=True)

    for line in _header_lines(position_set, exportable):
        sheet.append([line])
    sheet.append([])

    header_row = sheet.max_row + 1
    sheet.append([title for title, _ in _COLUMNS])
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in sheet[header_row]:
        cell.font = bold
        cell.fill = fill

    for index, position in enumerate(exportable, start=1):
        sheet.append(_position_row(index, position, formatted=False))

    sheet.append([])
    for unit, total in _totals(exportable):
        sheet.append(["", "", f"SUMME {unit}", round(total, 3), unit])
        for cell in sheet[sheet.max_row]:
            cell.font = bold

    for offset, (_, width) in enumerate(_COLUMNS):
        sheet.column_dimensions[chr(ord("A") + offset)].width = width
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ------------------------------------------------------------------------ csv

def build_csv(position_set: PositionSet) -> str:
    """Plain CSV for systems that cannot read xlsx. Same gate, same columns."""
    import csv
    from io import StringIO

    exportable = _gate(position_set)

    buffer = StringIO()
    # Semicolon: German Excel reads comma-separated files as a single column,
    # and our decimal separator is a comma anyway.
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow([title for title, _ in _COLUMNS])

    for index, position in enumerate(exportable, start=1):
        writer.writerow(_position_row(index, position, formatted=True))

    return buffer.getvalue()
