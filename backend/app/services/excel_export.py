"""
Excel Export Service

Generates Excel files from extraction results for German construction professionals.
Supports multiple export formats optimized for Aufmaß (measurement take-off) workflows.
"""

import io
from typing import Dict, List, Any, Optional
from datetime import datetime
from dataclasses import dataclass
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel export will not be available.")


@dataclass
class ExcelExportResult:
    """Result of Excel export."""
    success: bool
    filename: str
    file_bytes: Optional[bytes] = None
    error: Optional[str] = None
    row_count: int = 0


# Color scheme for German construction documents
COLORS = {
    "header_bg": "1F4E79",  # Dark blue
    "header_fg": "FFFFFF",  # White
    "category_bg": "D9E2F3",  # Light blue
    "total_bg": "FFC000",  # Gold
    "outdoor_bg": "E2EFDA",  # Light green (for balconies)
}


def is_excel_available() -> bool:
    """Check if Excel export is available."""
    return OPENPYXL_AVAILABLE


def export_extraction_to_excel(
    extraction_data: Dict[str, Any],
    source_filename: str = "blueprint.pdf",
    include_summary_sheet: bool = True,
    include_details_sheet: bool = True,
    include_category_sheets: bool = False,
    language: str = "de",
) -> ExcelExportResult:
    """
    Export extraction results to Excel file.

    Args:
        extraction_data: Dictionary with extraction results
        source_filename: Name of the source PDF file
        include_summary_sheet: Include summary overview sheet
        include_details_sheet: Include detailed room list
        include_category_sheets: Create separate sheets per category
        language: Output language (de/en)

    Returns:
        ExcelExportResult with file bytes
    """
    if not OPENPYXL_AVAILABLE:
        return ExcelExportResult(
            success=False,
            filename="",
            error="openpyxl not installed. Run: pip install openpyxl",
        )

    try:
        wb = openpyxl.Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        rooms = extraction_data.get("rooms", [])
        totals = extraction_data.get("totals_by_category", {})
        total_area = extraction_data.get("total_area_m2", 0)
        total_counted = extraction_data.get("total_counted_m2", 0)
        room_count = extraction_data.get("room_count", len(rooms))
        style = extraction_data.get("blueprint_style", "unknown")

        # Create summary sheet
        if include_summary_sheet:
            _create_summary_sheet(
                wb, rooms, totals, total_area, total_counted,
                room_count, style, source_filename, language
            )

        # Create details sheet
        if include_details_sheet:
            _create_details_sheet(wb, rooms, language)

        # Create category sheets
        if include_category_sheets:
            _create_category_sheets(wb, rooms, language)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = source_filename.rsplit(".", 1)[0] if "." in source_filename else source_filename
        excel_filename = f"Aufmass_{base_name}_{timestamp}.xlsx"

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return ExcelExportResult(
            success=True,
            filename=excel_filename,
            file_bytes=buffer.getvalue(),
            row_count=len(rooms),
        )

    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return ExcelExportResult(
            success=False,
            filename="",
            error=str(e),
        )


def _create_summary_sheet(
    wb: "openpyxl.Workbook",
    rooms: List[Dict],
    totals: Dict[str, float],
    total_area: float,
    total_counted: float,
    room_count: int,
    style: str,
    source_filename: str,
    language: str,
):
    """Create summary overview sheet."""
    ws = wb.create_sheet("Zusammenfassung" if language == "de" else "Summary")

    # Styles
    header_font = Font(bold=True, color=COLORS["header_fg"], size=12)
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    total_fill = PatternFill(start_color=COLORS["total_bg"], end_color=COLORS["total_bg"], fill_type="solid")
    bold_font = Font(bold=True)

    # Title
    ws["A1"] = "AUFMASS - Flächenextraktion" if language == "de" else "MEASUREMENT - Area Extraction"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    # Metadata
    ws["A3"] = "Quelldatei:" if language == "de" else "Source File:"
    ws["B3"] = source_filename
    ws["A4"] = "Extraktionsdatum:" if language == "de" else "Extraction Date:"
    ws["B4"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws["A5"] = "Grundrisstyp:" if language == "de" else "Blueprint Style:"
    ws["B5"] = style

    # Summary table header
    row = 8
    ws[f"A{row}"] = "Kategorie" if language == "de" else "Category"
    ws[f"B{row}"] = "Fläche (m²)" if language == "de" else "Area (m²)"
    ws[f"C{row}"] = "Räume" if language == "de" else "Rooms"

    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = header_font
        ws[f"{col}{row}"].fill = header_fill
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")

    # Category rows
    row += 1
    category_names_de = {
        "office": "Büro",
        "residential": "Wohnen",
        "circulation": "Verkehr",
        "stairs": "Treppen",
        "elevators": "Aufzüge",
        "shafts": "Schächte",
        "technical": "Technik",
        "sanitary": "Sanitär",
        "storage": "Lager",
        "outdoor": "Außen",
        "other": "Sonstige",
    }

    for cat, area in sorted(totals.items(), key=lambda x: -x[1]):
        cat_rooms = len([r for r in rooms if r.get("category") == cat])
        ws[f"A{row}"] = category_names_de.get(cat, cat) if language == "de" else cat.title()
        ws[f"B{row}"] = round(area, 2)
        ws[f"B{row}"].number_format = "#,##0.00"
        ws[f"C{row}"] = cat_rooms
        row += 1

    # Total row
    ws[f"A{row}"] = "GESAMT" if language == "de" else "TOTAL"
    ws[f"B{row}"] = round(total_area, 2)
    ws[f"B{row}"].number_format = "#,##0.00"
    ws[f"C{row}"] = room_count

    for col in ["A", "B", "C"]:
        ws[f"{col}{row}"].font = bold_font
        ws[f"{col}{row}"].fill = total_fill

    # Counted area (if different from total)
    if abs(total_counted - total_area) > 0.01:
        row += 2
        ws[f"A{row}"] = "Angerechnete Fläche:" if language == "de" else "Counted Area:"
        ws[f"B{row}"] = round(total_counted, 2)
        ws[f"B{row}"].number_format = "#,##0.00"
        ws[f"A{row}"].font = bold_font
        row += 1
        ws[f"A{row}"] = "(nach Balkon-Faktor 50%)" if language == "de" else "(after 50% balcony factor)"

    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10


def _create_details_sheet(
    wb: "openpyxl.Workbook",
    rooms: List[Dict],
    language: str,
):
    """Create detailed room list sheet."""
    ws = wb.create_sheet("Raumliste" if language == "de" else "Room List")

    # Styles
    header_font = Font(bold=True, color=COLORS["header_fg"])
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    outdoor_fill = PatternFill(start_color=COLORS["outdoor_bg"], end_color=COLORS["outdoor_bg"], fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    if language == "de":
        headers = ["Nr.", "Raumnummer", "Raumname", "Kategorie", "Fläche (m²)", "Faktor", "Angerechnet (m²)", "Seite", "Quelle"]
    else:
        headers = ["No.", "Room Number", "Room Name", "Category", "Area (m²)", "Factor", "Counted (m²)", "Page", "Source"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for row_idx, room in enumerate(rooms, 2):
        is_outdoor = room.get("category") == "outdoor"

        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=room.get("room_number", ""))
        ws.cell(row=row_idx, column=3, value=room.get("room_name", ""))
        ws.cell(row=row_idx, column=4, value=room.get("category", ""))
        ws.cell(row=row_idx, column=5, value=room.get("area_m2", 0))
        ws.cell(row=row_idx, column=6, value=room.get("factor", 1.0))
        ws.cell(row=row_idx, column=7, value=room.get("counted_m2", room.get("area_m2", 0)))
        ws.cell(row=row_idx, column=8, value=room.get("page", 0))
        ws.cell(row=row_idx, column=9, value=room.get("source_text", ""))

        # Format numbers
        ws.cell(row=row_idx, column=5).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=6).number_format = "0.00"
        ws.cell(row=row_idx, column=7).number_format = "#,##0.00"

        # Apply border and highlight outdoor rows
        for col in range(1, 10):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if is_outdoor:
                cell.fill = outdoor_fill

    # Adjust column widths
    column_widths = [5, 15, 25, 15, 12, 8, 12, 6, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Add total row
    total_row = len(rooms) + 2
    ws.cell(row=total_row, column=4, value="GESAMT:" if language == "de" else "TOTAL:")
    ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row-1})")

    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = "#,##0.00"
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = "#,##0.00"

    # Freeze header row
    ws.freeze_panes = "A2"


def _create_category_sheets(
    wb: "openpyxl.Workbook",
    rooms: List[Dict],
    language: str,
):
    """Create separate sheets for each category."""
    # Group rooms by category
    by_category: Dict[str, List[Dict]] = {}
    for room in rooms:
        cat = room.get("category", "other")
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(room)

    category_names_de = {
        "office": "Büro",
        "residential": "Wohnen",
        "circulation": "Verkehr",
        "stairs": "Treppen",
        "elevators": "Aufzüge",
        "shafts": "Schächte",
        "technical": "Technik",
        "sanitary": "Sanitär",
        "storage": "Lager",
        "outdoor": "Außen",
        "other": "Sonstige",
    }

    for cat, cat_rooms in by_category.items():
        sheet_name = category_names_de.get(cat, cat)[:31] if language == "de" else cat.title()[:31]
        ws = wb.create_sheet(sheet_name)

        # Simple list for category
        headers = ["Raumnummer", "Raumname", "Fläche (m²)"] if language == "de" else ["Room Number", "Room Name", "Area (m²)"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row_idx, room in enumerate(cat_rooms, 2):
            ws.cell(row=row_idx, column=1, value=room.get("room_number", ""))
            ws.cell(row=row_idx, column=2, value=room.get("room_name", ""))
            ws.cell(row=row_idx, column=3, value=room.get("area_m2", 0))
            ws.cell(row=row_idx, column=3).number_format = "#,##0.00"

        # Total
        total_row = len(cat_rooms) + 2
        ws.cell(row=total_row, column=2, value="Gesamt:" if language == "de" else "Total:")
        ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=3).number_format = "#,##0.00"

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 12


def export_to_csv(
    extraction_data: Dict[str, Any],
    source_filename: str = "blueprint.pdf",
    language: str = "de",
) -> ExcelExportResult:
    """
    Export extraction results to CSV format.

    Fallback when Excel is not available.
    """
    import csv

    try:
        rooms = extraction_data.get("rooms", [])

        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";")  # German Excel uses semicolon

        # Headers
        if language == "de":
            headers = ["Raumnummer", "Raumname", "Kategorie", "Fläche (m²)", "Faktor", "Angerechnet (m²)", "Seite", "Quelle"]
        else:
            headers = ["Room Number", "Room Name", "Category", "Area (m²)", "Factor", "Counted (m²)", "Page", "Source"]

        writer.writerow(headers)

        # Data
        for room in rooms:
            writer.writerow([
                room.get("room_number", ""),
                room.get("room_name", ""),
                room.get("category", ""),
                str(room.get("area_m2", 0)).replace(".", ","),  # German decimal
                str(room.get("factor", 1.0)).replace(".", ","),
                str(room.get("counted_m2", room.get("area_m2", 0))).replace(".", ","),
                room.get("page", 0),
                room.get("source_text", ""),
            ])

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = source_filename.rsplit(".", 1)[0] if "." in source_filename else source_filename
        csv_filename = f"Aufmass_{base_name}_{timestamp}.csv"

        return ExcelExportResult(
            success=True,
            filename=csv_filename,
            file_bytes=buffer.getvalue().encode("utf-8-sig"),  # BOM for Excel
            row_count=len(rooms),
        )

    except Exception as e:
        logger.error(f"CSV export error: {e}")
        return ExcelExportResult(
            success=False,
            filename="",
            error=str(e),
        )
